import csv
import json
from pathlib import Path
import pytest

from NanoVNASaver.TestSpec import TestPoint, TestResult, TestData
from NanoVNASaver.Controls.LotControl import LotControl


def _make_testdata():
    tp1 = TestPoint(name="TP One", parameter="S21", frequency=900000000, span=3000000, limit_db=-30.0, direction="under")
    tp2 = TestPoint(name="TP Two", parameter="S11", frequency=1700000000, span=1000000, limit_db=-6.0, direction="over")
    r1 = TestResult(tp=tp1, passed=True, min=-29.5, max=-25.0, failing=[], samples=10)
    r2 = TestResult(tp=tp2, passed=False, min=-8.0, max=0.0, failing=[1700000000], samples=5)
    td = TestData(serial="SN123", id="tid-1", meta="m", passed=False, pcb_lot="LOT42", results=[r1, r2])
    return td


def test_log_write_csv(tmp_path: Path):
    td = _make_testdata()
    csv_path = tmp_path / "lotA_log.csv"
    # Call as unbound method (no UI instance required)
    LotControl.log_write(None, td, csv_path, filter=None, excel=False)

    assert csv_path.exists()
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    assert len(rows) == 1
    row = rows[0]
    assert row["serial"] == "SN123"
    assert row["id"] == "tid-1"
    # header should include prefixed fields for both test points
    assert any(k.startswith("TP_One_") or k.startswith("TP_One") for k in row.keys())
    assert any(k.startswith("TP_Two_") or k.startswith("TP_Two") for k in row.keys())


def test_log_write_csv_with_filter(tmp_path: Path):
    td = _make_testdata()
    csv_path = tmp_path / "lotB_log.csv"
    # Filter should only apply to per-test fields; top-level fields remain present
    LotControl.log_write(None, td, csv_path, filter=["passed", "limit_db"], excel=False)
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
    # ensure top-level passed is still present
    assert "passed" in header
    # ensure limit_db columns (prefixed) are not present
    assert not any("limit_db" in h for h in header)
    # ensure per-test 'passed' columns were filtered (top-level 'passed' remains)
    assert not any(h.endswith("_passed") for h in header)


def test_log_write_excel(tmp_path: Path):
    pytest.importorskip("openpyxl")
    td = _make_testdata()
    xlsx_path = tmp_path / "lotC_log.xlsx"
    LotControl.log_write(None, td, xlsx_path, filter=None, excel=True)
    assert xlsx_path.exists()

    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    assert "serial" in header
    # first data row
    data_row = [cell.value for cell in ws[2]]
    # find serial column
    si = header.index("serial")
    assert data_row[si] == "SN123"
