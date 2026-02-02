#  NanoVNASaver
#
#  A python program to view and export Touchstone data from a NanoVNA
#  Copyright (C) 2019, 2020  Rune B. Broberg
#  Copyright (C) 2020,2021 NanoVNA-Saver Authors
#
#  This program is free software: you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation, either version 3 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program.  If not, see <https://www.gnu.org/licenses/>.

import logging
from pathlib import Path
from statistics import mode
from typing import TYPE_CHECKING, Optional
import json
from datetime import datetime
from ..Touchstone import Touchstone
from ..TestSpec import TestResult, TestData
from PySide6 import QtCore, QtWidgets
from PySide6.QtCore import Signal, Qt, QUrl
from PySide6.QtGui import QDesktopServices

from .Control import Control

if TYPE_CHECKING:
    from ..NanoVNASaver.NanoVNASaver import NanoVNASaver as vna_app


logger = logging.getLogger(__name__)


class LotControl(Control):
    """Dummy Lot control with a path field, serial label and a scrollable list of serials and statuses."""

    lot_changed = Signal(bool)
    # Emitted when the PCB lot is explicitly set via the 'Set' button
    pcb_lot_changed = Signal()

    def __init__(self, app: "vna_app"):
        super().__init__(app, "Lot control")

        # Internal state
        self.current_lot_path: Path | None = None
        self.current_lot_name: str | None = None
        self.lot_counter = 1
        self.lots: dict[str, str] = {}
        self.lot_samples: dict[str, int] = {}
        self.lot_passed: dict[str, int] = {}
        self.lot_failed: dict[str, int] = {}
        # Per-lot unit tracking: list of [serial, passed(bool)]
        # Stored on disk as a JSON list of pairs, e.g. [["SN1", true], ["SN2", false]]
        self.lot_units: dict[str, list] = {}
        # Counts derived from lot_units
        self.lot_passed_units: dict[str, int] = {}
        self.lot_failed_units: dict[str, int] = {}
        # Per-lot test spec checksum (value of SweepEvaluate.test_checksum at creation)
        self.lot_checksum: dict[str, str | None] = {}
        self.highlighted_row: int | None = None
        # PCB lot field state: True if the PCB Lot text field is empty
        self.pcb_lot_empty: bool = True
        # Stored PCB lot value set via the 'Set' button; displayed in the PCB lot indicator
        self.pcb_lot_value: str | None = None

        # Working directory for new lots / defaults
        self.working_directory: Path = Path.cwd()

        # Lot label: plain text showing current lot (default "No lot")
        self.lot_label = QtWidgets.QLabel("No lot")
        self.lot_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.lot_label.setMinimumHeight(24)
        self.lot_label.setAlignment(Qt.AlignCenter)
        #self.golden_label.setMinimumHeight(34)
        self.lot_label.setStyleSheet(
            "padding: 1px; border-radius: 4px; border: 2px solid #666; font-size: 18px;"
        )
        self.layout.addRow(QtWidgets.QLabel("Current lot"))
        self.layout.addRow(self.lot_label)  # spacer

        # Yield label: percentage display for the currently selected lot (same styling as current lot)
        self.yield_label = QtWidgets.QLabel("N/A")
        self.yield_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.yield_label.setMinimumHeight(24)
        self.yield_label.setAlignment(Qt.AlignCenter)
        self.yield_label.setStyleSheet(
            "padding: 1px; border-radius: 4px; border: 2px solid #666; font-size: 18px;"
        )
        self.layout.addRow(QtWidgets.QLabel("Yield"))
        self.layout.addRow(self.yield_label)

        # PCB Lot indicator: shows the currently set PCB lot (via the Set button)
        self.pcb_lot_indicator = QtWidgets.QLabel("N/A")
        self.pcb_lot_indicator.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.pcb_lot_indicator.setMinimumHeight(24)
        self.pcb_lot_indicator.setAlignment(Qt.AlignCenter)
        self.pcb_lot_indicator.setStyleSheet(
            "padding: 1px; border-radius: 4px; border: 2px solid #666; font-size: 18px;"
        )
        self.layout.addRow(QtWidgets.QLabel("PCB Lot"))
        self.layout.addRow(self.pcb_lot_indicator)

        # Samples label: plain numeric display for the currently selected lot
        self.samples_label = QtWidgets.QLabel("0")
        self.samples_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.samples_label.setMinimumHeight(24)
        self.layout.addRow("Samples:", self.samples_label)

        # Passed / Failed counts for the currently selected lot
        self.passed_label = QtWidgets.QLabel("0")
        self.passed_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.passed_label.setMinimumHeight(24)
        self.layout.addRow("Passed count:", self.passed_label)

        self.failed_label = QtWidgets.QLabel("0")
        self.failed_label.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)
        self.failed_label.setMinimumHeight(24)
        self.layout.addRow("Failed count:", self.failed_label)

        # Current serial label
        #self.serial_label = QtWidgets.QLabel("#N/A")
        #self.layout.addRow("Current serial:", self.serial_label)

        # PCB Lot input field (with 'Set' button to store a persistent PCB lot indicator)
        self.pcb_lot_field = QtWidgets.QLineEdit()
        self.pcb_lot_field.setPlaceholderText("Enter PCB lot number")
        self.pcb_lot_field.setMinimumHeight(24)
        # Row with input + Set button
        pcb_row = QtWidgets.QWidget()
        pcb_row_layout = QtWidgets.QHBoxLayout()
        pcb_row_layout.setContentsMargins(0, 0, 0, 0)
        pcb_row.setLayout(pcb_row_layout)
        pcb_row_layout.addWidget(self.pcb_lot_field)
        self.btn_set_pcb_lot = QtWidgets.QPushButton("Set")
        self.btn_set_pcb_lot.setMinimumHeight(24)
        pcb_row_layout.addWidget(self.btn_set_pcb_lot)
        self.layout.addRow("PCB Lot:", pcb_row)
        # Ensure internal state tracks the field (default is empty)
        self.pcb_lot_empty = True
        self.pcb_lot_field.textChanged.connect(self._on_pcb_lot_changed)
        self.btn_set_pcb_lot.clicked.connect(self._on_set_pcb_lot) 

        # Table for lots (Lot name + number of samples)
        self.table = QtWidgets.QTableWidget(0, 2)
        self.table.setHorizontalHeaderLabels(["Lot", "Samples"])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setMinimumHeight(150)
        self.layout.addRow(self.table)

        # Buttons
        btn_layout = QtWidgets.QHBoxLayout()

        self.btn_select = QtWidgets.QPushButton("Select")
        self.btn_select.setMinimumHeight(20)
        self.btn_select.setEnabled(False)
        self.btn_select.clicked.connect(self.select_lot)
        btn_layout.addWidget(self.btn_select)

        self.btn_new = QtWidgets.QPushButton("New lot")
        self.btn_new.setMinimumHeight(20)
        self.btn_new.clicked.connect(self.new_lot)
        btn_layout.addWidget(self.btn_new)

        self.btn_open = QtWidgets.QPushButton("Directory")
        self.btn_open.setMinimumHeight(20)
        # Directory now prompts to change the working directory; always enabled
        self.btn_open.setEnabled(True)
        self.btn_open.clicked.connect(self.open_directory)
        btn_layout.addWidget(self.btn_open)

        self.layout.addRow(btn_layout)

        

        # Connect click handler (highlighting). Selection will be performed with the Select button
        self.table.itemClicked.connect(self._on_item_clicked)

        # Scan the working directory for existing lot directories
        try:
            self.scan_working_directory()
        except Exception:
            logger.exception("Failed scanning working directory")

    def set_lot_selected(self, lot_name: str, path: str) -> None:
        """Set the selected lot and update UI state."""
        self.current_lot_name = lot_name
        self.current_lot_path = Path(path)
        self.lot_label.setText(lot_name)
        # Update samples display for selected lot
        samples = self.lot_samples.get(lot_name, 0)
        self.samples_label.setText(f"{samples}")
        passed = self.lot_passed.get(lot_name, 0)
        failed = self.lot_failed.get(lot_name, 0)
        self.passed_label.setText(str(passed))
        self.failed_label.setText(str(failed))
        # Compute and display yield as percentage with two decimals
        try:
            total = passed + failed
            yield_pct = (float(passed) / total * 100.0) if total else 0.0
        except Exception:
            yield_pct = 0.0
        if hasattr(self, "yield_label"):
            self.yield_label.setText(f"{yield_pct:.2f}%")
        #self.serial_label.setText("#N/A")
        # Preserve PCB lot indicator and input when selecting a different lot
        # (Do not clear PCB lot set by the user)
        self.btn_open.setEnabled(True)
        self.btn_select.setEnabled(False)
        # visually select the row
        for r in range(self.table.rowCount()):
            if self.table.item(r, 0).text() == lot_name:
                self.table.selectRow(r)
                break
        self.lot_changed.emit(True)

    def new_lot(self) -> None:
        """Prompt for a new lot name and add it to the list."""
        default_name = f"lot_{self.lot_counter:03d}"
        name, ok = QtWidgets.QInputDialog.getText(
            self, "New lot", "Lot name:", QtWidgets.QLineEdit.Normal, default_name
        )
        if not ok or not name:
            return
        self.lot_counter += 1
        # create on disk in working_directory
        self.add_lot(name, None, samples=0, create_on_disk=True)
        # Highlight the newly added lot (does not yet select it)
        for r in range(self.table.rowCount()):
            if self.table.item(r, 0).text() == name:
                self.table.setCurrentCell(r, 0)
                self.highlighted_row = r
                self.btn_select.setEnabled(True)
                break

    def select_lot(self) -> None:
        """Select the currently highlighted lot (finalize selection)."""
        if self.highlighted_row is None:
            return
        lot_name = self.table.item(self.highlighted_row, 0).text()
        lot_path = self.lots.get(lot_name, str(self.working_directory / lot_name))
        # refresh samples from disk if present
        lot_dir = Path(lot_path)
        info_file = lot_dir / f"{lot_name}.json"
        if info_file.exists():
            try:
                with info_file.open("r", encoding="utf-8") as f:
                    info = json.load(f)
                samples = int(info.get("samples", 0))
                units = info.get("units", []) if isinstance(info.get("units", []), list) else []
                passed_units = int(info.get("passed_units", info.get("passed", 0)))
                failed_units = int(info.get("failed_units", info.get("failed", 0)))
                if units and ("passed_units" not in info or "failed_units" not in info):
                    p = sum(1 for u in units if bool(u[1]))
                    f = len(units) - p
                    passed_units = p
                    failed_units = f
                self.lot_samples[lot_name] = samples
                # Keep legacy names for UI compatibility
                self.lot_passed[lot_name] = int(passed_units)
                self.lot_failed[lot_name] = int(failed_units)
                # Store new structures
                self.lot_passed_units[lot_name] = int(passed_units)
                self.lot_failed_units[lot_name] = int(failed_units)
                self.lot_units[lot_name] = list(units)
                # preserve checksum from disk if present
                self.lot_checksum[lot_name] = info.get("checksum", None)
            except Exception:
                logger.exception("Failed reading lot info for %s", lot_name)
        self.set_lot_selected(lot_name, lot_path)

    def open_directory(self) -> None:
        """Prompt the user to change the working directory for lots and scan it."""
        dir_path = QtWidgets.QFileDialog.getExistingDirectory(
            self, "Select working directory", str(self.working_directory)
        )
        if not dir_path:
            return
        self.working_directory = Path(dir_path)
        logger.info("Working directory changed to %s", self.working_directory)
        # Rescan working directory for lot directories
        try:
            self.scan_working_directory()
        except Exception:
            logger.exception("Failed scanning new working directory")
        QtWidgets.QMessageBox.information(
            self, "Working directory",
            f"Working directory set to:\n{self.working_directory}"
        )

    def scan_working_directory(self) -> None:
        """Scan the working directory for lot directories and populate the table.

        A lot directory is a subdirectory that contains a JSON file with keys:
        - "lot_name" (must match the directory name)
        - "samples" (int)
        - "creation_date" (ISO datetime string)
        """
        # Clear current lists and table
        self.lots.clear()
        self.lot_samples.clear()
        self.table.setRowCount(0)

        if not self.working_directory.exists():
            return

        for child in sorted(self.working_directory.iterdir()):
            if not child.is_dir():
                continue
            # look for json files inside child
            info_file = None
            for candidate in child.glob("*.json"):
                try:
                    with candidate.open("r", encoding="utf-8") as f:
                        info = json.load(f)
                    # validate
                    if (
                        isinstance(info, dict)
                        and info.get("lot_name") == child.name
                        and "samples" in info
                        and "creation_date" in info
                    ):
                        info_file = candidate
                        samples = int(info.get("samples", 0))
                        # Units list stored as list of [serial, passed_bool]
                        units = info.get("units", []) if isinstance(info.get("units", []), list) else []
                        # Prefer explicit passed_units/failed_units when available, else fall back to legacy fields
                        passed_units = int(info.get("passed_units", info.get("passed", 0)))
                        failed_units = int(info.get("failed_units", info.get("failed", 0)))
                        # If units present and explicit counts not provided, infer counts from units
                        if units and ("passed_units" not in info or "failed_units" not in info):
                            p = sum(1 for u in units if bool(u[1]))
                            f = len(units) - p
                            passed_units = p
                            failed_units = f
                        # preserve checksum from disk if present
                        self.lot_checksum[child.name] = info.get("checksum", None)
                        # add lot without creating on disk
                        self.add_lot(child.name, str(child), samples=samples, passed_units=passed_units, failed_units=failed_units, units=units, create_on_disk=False)
                        break
                except Exception:
                    logger.exception("Invalid lot json at %s", candidate)
                    continue

    def add_lot(self, lot_name: str, path: str | None = None, *, samples: int = 0, passed_units: int = 0, failed_units: int = 0, units: list | None = None, create_on_disk: bool = True) -> None:
        """Add a lot entry to the table (name + sample count).

        If create_on_disk is True, this will create a directory in self.working_directory
        and write a `{lot_name}.json` file with the lot metadata (if it does not already exist).

        New JSON fields supported:
        - "passed_units": number of uniquely passing units (int)
        - "failed_units": number of uniquely failing units (int)
        - "units": list of [serial, passed_bool] entries
        - "yield": float (passed_units / len(units), 0.0 when no units)
        """
        # Determine directory
        lot_dir = Path(path) if path else (self.working_directory / lot_name)
        if units is None:
            units = []
        if create_on_disk:
            try:
                lot_dir.mkdir(parents=True, exist_ok=True)
                info_file = lot_dir / f"{lot_name}.json"
                if not info_file.exists():
                    # Determine checksum from SweepEvaluate (if available)
                    try:
                        se = getattr(self.app, "sweep_evaluate", None)
                        checksum_val = getattr(se, "test_checksum", None)
                    except Exception:
                        checksum_val = None

                    info = {
                        "lot_name": lot_name,
                        "samples": int(samples),
                        # Use new field names for unit counts
                        "passed_units": int(passed_units),
                        "failed_units": int(failed_units),
                        "units": list(units),
                        "yield": (float(passed_units) / len(units)) if len(units) else 0.0,
                        "checksum": checksum_val,
                        "creation_date": datetime.now().isoformat(),
                    }
                    with info_file.open("w", encoding="utf-8") as f:
                        json.dump(info, f, indent=2)
                    # keep internal state for checksum
                    self.lot_checksum[lot_name] = checksum_val
                else:
                    # read existing info
                    try:
                        with info_file.open("r", encoding="utf-8") as f:
                            existing = json.load(f)
                        samples = int(existing.get("samples", samples))
                        # Prefer new field names when present
                        passed_units = int(existing.get("passed_units", existing.get("passed", passed_units)))
                        failed_units = int(existing.get("failed_units", existing.get("failed", failed_units)))
                        units = existing.get("units", units) or units
                        # preserve checksum from existing info if present
                        self.lot_checksum[lot_name] = existing.get("checksum", None)
                    except Exception:
                        logger.exception("Failed reading existing lot info %s", info_file)
            except Exception:
                logger.exception("Failed to create lot directory %s", lot_dir)

        # If already present, update sample count and path
        if lot_name in self.lots:
            self.lots[lot_name] = str(lot_dir)
            self.lot_samples[lot_name] = int(samples)
            # keep compatibility with existing naming used by UI
            self.lot_passed[lot_name] = int(passed_units)
            self.lot_failed[lot_name] = int(failed_units)
            # store units list
            self.lot_units[lot_name] = list(units)
            self.lot_passed_units[lot_name] = int(passed_units)
            self.lot_failed_units[lot_name] = int(failed_units)
            # update table row for samples
            for r in range(self.table.rowCount()):
                if self.table.item(r, 0).text() == lot_name:
                    self.table.item(r, 1).setText(str(self.lot_samples[lot_name]))
                    break
            return

        # otherwise insert new row
        row = self.table.rowCount()
        self.table.insertRow(row)
        name_item = QtWidgets.QTableWidgetItem(lot_name)
        name_item.setFlags(name_item.flags() ^ QtCore.Qt.ItemIsEditable)
        self.table.setItem(row, 0, name_item)

        self.lots[lot_name] = str(lot_dir)
        self.lot_samples[lot_name] = int(samples)
        # keep compatibility with existing naming used by UI
        self.lot_passed[lot_name] = int(passed_units)
        self.lot_failed[lot_name] = int(failed_units)
        # store units list
        self.lot_units[lot_name] = list(units)
        self.lot_passed_units[lot_name] = int(passed_units)
        self.lot_failed_units[lot_name] = int(failed_units)
        samples_item = QtWidgets.QTableWidgetItem(str(self.lot_samples[lot_name]))
        samples_item.setFlags(samples_item.flags() ^ QtCore.Qt.ItemIsEditable)
        self.table.setItem(row, 1, samples_item)

    def _on_pcb_lot_changed(self, text: str) -> None:
        """Update internal state tracking whether PCB Lot field is empty."""
        self.pcb_lot_empty = (text.strip() == "")

    def _on_set_pcb_lot(self) -> None:
        """Set PCB lot indicator from the input field and preserve it across lot changes."""
        try:
            text = self.pcb_lot_field.text().strip()
            if not text:
                QtWidgets.QMessageBox.warning(self, "Invalid PCB lot", "No PCB lot entered")
                return
            self.pcb_lot_value = text
            if hasattr(self, "pcb_lot_indicator"):
                self.pcb_lot_indicator.setText(text)
            # Mark input as non-empty now that a value has been set
            self.pcb_lot_empty = False
            # Notify listeners that a PCB lot has been explicitly set
            try:
                self.pcb_lot_changed.emit()
            except Exception:
                pass
        except Exception:
            logger.exception("Failed setting PCB lot")

    def _on_item_clicked(self, item: QtWidgets.QTableWidgetItem) -> None:
        # store the highlighted row, but do not finalize selection until user presses Select
        self.highlighted_row = item.row()
        self.btn_select.setEnabled(True)
        
    def save_results_for_latest(self, test_data):
        print("LotControl: save_results_for_latest called")
        """Save TestData into the currently selected lot directory.

        For each TestResult in the TestData this saves:
        - one S1P covering the test frequency range (if S11 samples available)
        - one S2P covering the test frequency range (if S11 and S21 samples available)
        and also saves `results_<serial>_<id>.json` containing the TestData contents.

        Files are written to: current_lot_path/<serial>/<id>/
        """
        # Accept legacy input (list of TestResult) for backward compatibility
        if not test_data:
            print("LotControl: No test data to save")
            return
        if not self.current_lot_path:
            QtWidgets.QMessageBox.information(self, "No lot selected", "Please select a lot first.")
            return
        out_dir = Path(self.current_lot_path)

        # Try to accept either a TestData object or a legacy list of TestResult
        try:
            dtn = datetime.now()
            ts = dtn.isoformat()
            results = getattr(test_data, "results", None)
            serial = getattr(test_data, "serial", None)
            print(f"LotControl: Sarjanumero {serial}")
            tid = getattr(test_data, "id", None)
            passed = getattr(test_data, "passed", None)
            # Prefer an explicitly set PCB Lot (via the 'Set' button), fall back to input field, then test_data
            pcb_lot = None
            try:
                if getattr(self, "pcb_lot_value", None):
                    pcb_lot = self.pcb_lot_value
                elif hasattr(self, "pcb_lot_field") and self.pcb_lot_field.text().strip():
                    pcb_lot = self.pcb_lot_field.text().strip()
                else:
                    pcb_lot = getattr(test_data, "pcb_lot", None)
            except Exception:
                pcb_lot = getattr(test_data, "pcb_lot", None)
            meta = getattr(test_data, "meta", None)
            # Ensure the TestData instance carries the PCB lot value we computed so that
            # subsequent writers (CSV/Excel) see the same value as we put in the results JSON.
            try:
                setattr(test_data, "pcb_lot", pcb_lot)
                setattr(test_data, "timestamp", ts)
            except Exception:
                # test_data might be a plain list (legacy mode) or otherwise not allow attributes
                pass

            if results is None and isinstance(test_data, list):
                # Legacy list: wrap into a simple TestData-like struct
                results = test_data
                serial = getattr(self, "current_lot_name", None) or "NOSERIAL"
                from uuid import uuid4
                tid = str(uuid4())
                passed = "UNKNOWN"
                pcb_lot = "UNKNOWN"
                meta = "legacy"
                
        except Exception:
            logger.exception("Invalid test_data passed to save_results_for_latest")
            return

        ts_full = getattr(self.app, "data", None)
        if ts_full is None:
            logger.warning("No sweep data available to save for TestData %s", tid)
            return

        saved_any = False
        # create sample dir for this TestData
        serial_dir = str(serial) if serial else str(tid)
        ts_label = dtn.strftime("%d_%m_%y_%H-%M")
        sample_dir = out_dir / serial_dir / f"{ts_label}_{tid}"
        print(f"LotControl: Saving TestData to {sample_dir}")
        sample_dir.mkdir(parents=True, exist_ok=True)

        # Save the TestData contents as JSON: results_<serial>_<id>.json
        results_meta_file = sample_dir / f"results_{serial}_{tid}.json"
        results_table_file = sample_dir / f"table_{serial}_{tid}.csv"
        results_table = []
        
        try:
            # Build a JSON-friendly representation
            # Determine test spec checksum from SweepEvaluate if available
            try:
                se = getattr(self.app, "sweep_evaluate", None)
                test_checksum_val = getattr(se, "test_checksum", None) if se is not None else None
            except Exception:
                test_checksum_val = None

            jt = {
                "serial": serial,
                "id": tid,
                "timestamp": ts,
                "meta": meta,
                "passed": passed,
                "pcb_lot": pcb_lot,
                "test_checksum": test_checksum_val,
                "results": [],
            }
            # Also attach checksum to the test_data instance where possible so CSV writer can access it
            try:
                setattr(test_data, "test_checksum", test_checksum_val)
            except Exception:
                pass
        
            for r in results:
                tp = getattr(r, "tp", None)
                tp_dict = None
                table_dict={}
                if tp is not None:
                    tp_dict = {
                        "name": getattr(tp, "name", None),
                        "parameter": getattr(tp, "parameter", None),
                        "frequency": getattr(tp, "frequency", None),
                        "span": getattr(tp, "span", None),
                        "limit_db": getattr(tp, "limit_db", None),
                        "direction": getattr(tp, "direction", None),
                    }

                table_dict = {
                    "passed": getattr(r, "passed", None),
                    "min": getattr(r, "min", None),
                    "max": getattr(r, "max", None),
                    "samples": getattr(r, "samples", None),
                }

                table_dict = {**table_dict, **tp_dict} if tp_dict else table_dict
                results_table.append(table_dict)

                jt["results"].append({
                    "tp": tp_dict,
                    "passed": getattr(r, "passed", None),
                    "min": getattr(r, "min", None),
                    "max": getattr(r, "max", None),
                    "failing": getattr(r, "failing", None),
                    "samples": getattr(r, "samples", None),
                })
            
            import csv as _csv
            try:
                p = Path(results_table_file)
                fieldnames = ["name", "parameter", "frequency", "span", "limit_db", "direction", "passed", "min", "max", "samples"]
                with p.open("w", encoding="utf-8", newline="") as f:
                    writer = _csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writeheader()
                    writer.writerows(results_table)
            except Exception:
                logger.exception("Failed writing result CSV table to %s", p)
         
            with results_meta_file.open("w", encoding="utf-8") as f:
                json.dump(jt, f, indent=2)
            saved_any = True

        except Exception:
            logger.exception("Failed saving TestData JSON %s", results_meta_file)

        # Save full Touchstone data: ONE S1P and ONE S2P per TestData (no per-test aggregation)
        from ..Windows.Files import FilesWindow

        try:
            #saved_s1 = FilesWindow.exportFileToDir(self, str(sample_dir), filename=f"{serial}_{tid}_s1p.s1p", nr_params=1)
            #logger.info("Saved full S1P for TestData %s to %s", tid, saved_s1)
            saved_s2 = FilesWindow.exportFileToDir(self, str(sample_dir), filename=f"{serial}_{tid}_s2p.s2p", nr_params=4)
            logger.info("Saved full S2P for TestData %s to %s", tid, saved_s2)

            # Save screenshots from SweepEvaluate charts (if available on the app)
            try:
                se = getattr(self.app, "sweep_evaluate", None)
                if se is not None:
                    # S11 chart screenshot
                    try:
                        s11_path = sample_dir / f"S11_{serial}_{tid}.png"
                        se.s11_chart.saveScreenshotTo(s11_path)
                        logger.info("Saved S11 screenshot to %s", s11_path)
                    except Exception:
                        logger.exception("Failed saving S11 screenshot for TestData %s", tid)
                    # S21 chart screenshot
                    try:
                        s21_path = sample_dir / f"S21_{serial}_{tid}.png"
                        se.s21_chart.saveScreenshotTo(s21_path)
                        logger.info("Saved S21 screenshot to %s", s21_path)
                    except Exception:
                        logger.exception("Failed saving S21 screenshot for TestData %s", tid)
            except Exception:
                logger.exception("Unexpected error while attempting to save chart screenshots for %s", tid)

        except Exception:
            logger.exception("Failed saving full S-parameter files for TestData %s", tid)
        # update sample count and lot.json (stored at lot root)
        lot_name = self.current_lot_name
        if lot_name and saved_any:
            self.lot_samples[lot_name] = self.lot_samples.get(lot_name, 0) + 1

            # Ensure unit tracking structures exist
            units = self.lot_units.get(lot_name, [])
            passed_units = self.lot_passed_units.get(lot_name, 0)
            failed_units = self.lot_failed_units.get(lot_name, 0)

            # Unit identifier: prefer serial, fallback to tid if no serial present
            unit_id = str(serial) if serial else str(tid)

            # Update units list and counts only when TestData.passed is a boolean
            if isinstance(passed, bool):
                # Find existing entry
                found_index = None
                for i, u in enumerate(units):
                    try:
                        if str(u[0]) == unit_id:
                            found_index = i
                            break
                    except Exception:
                        continue

                if found_index is None:
                    # New unit: append and update counts
                    units.append([unit_id, bool(passed)])
                    if passed:
                        passed_units += 1
                    else:
                        failed_units += 1
                else:
                    # Existing unit: check state transitions
                    existing_passed = bool(units[found_index][1])
                    if passed and not existing_passed:
                        # Failure -> pass transition
                        units[found_index][1] = True
                        passed_units += 1
                        failed_units = max(0, failed_units - 1)
                    elif not passed and existing_passed:
                        # Pass -> failure transition
                        units[found_index][1] = False
                        failed_units += 1
                        passed_units = max(0, passed_units - 1)
                    else:
                        # No state change
                        pass

            # Persist updated state back to internal structures
            self.lot_units[lot_name] = units
            self.lot_passed_units[lot_name] = passed_units
            self.lot_failed_units[lot_name] = failed_units
            # Keep compatibility fields used by UI
            self.lot_passed[lot_name] = passed_units
            self.lot_failed[lot_name] = failed_units

            # Compute yield: passed_units / total units
            try:
                yield_val = float(passed_units) / len(units) if len(units) else 0.0
            except Exception:
                yield_val = 0.0

            info_file = out_dir / f"{lot_name}.json"
            try:
                info = {
                    "lot_name": lot_name,
                    "samples": int(self.lot_samples[lot_name]),
                    "passed_units": int(self.lot_passed_units[lot_name]),
                    "failed_units": int(self.lot_failed_units[lot_name]),
                    "yield": float(yield_val),
                    "checksum": self.lot_checksum.get(lot_name, None),
                    "creation_date": datetime.now().isoformat(),
                    "units": list(self.lot_units.get(lot_name, [])),
                }
                with info_file.open("w", encoding="utf-8") as f:
                    json.dump(info, f, indent=2)    
            except Exception:
                logger.exception("Failed updating lot info %s", info_file)

        # refresh UI sample count and pass/fail counts
        if saved_any and self.current_lot_name:
            self.samples_label.setText(str(self.lot_samples.get(self.current_lot_name, 0)))
            self.passed_label.setText(str(self.lot_passed.get(self.current_lot_name, 0)))
            self.failed_label.setText(str(self.lot_failed.get(self.current_lot_name, 0)))
            # Update yield display
            try:
                passed = self.lot_passed.get(self.current_lot_name, 0)
                failed = self.lot_failed.get(self.current_lot_name, 0)
                total = passed + failed
                yield_pct = (float(passed) / total * 100.0) if total else 0.0
            except Exception:
                yield_pct = 0.0
            if hasattr(self, "yield_label"):
                self.yield_label.setText(f"{yield_pct:.2f}%")

        # Append a CSV log entry for this test data to a per-lot CSV file
        try:
            if lot_name and saved_any:
                csv_path = out_dir / f"{lot_name}_log.csv"
                # Default: no filter, write CSV (excel=False)
                try:
                    self.log_write(test_data, csv_path, filter=None, excel=False)
                    self.log_write(test_data, csv_path, ["samples","failing","passed"], excel=True) # Also write a human-readable Excel version
                except Exception:
                    logger.exception("Failed writing CSV log for %s to %s", tid, csv_path)
        except Exception:
            logger.exception("Unexpected error when attempting to write CSV log")

    def log_write(self, test_data: TestData, path: Path, filter: list[str] | None = None, excel: bool = False) -> None:
        """Write a single CSV or Excel row for the provided TestData to the file at ``path``.

        Parameters
        - test_data: TestData to log
        - path: Path to CSV or XLSX file (if ``excel=True`` an ``.xlsx`` file will be used)
        - filter: list of attribute names to omit (checks un-prefixed names like "passed", "limit_db")
        - excel: if True, write an Excel (.xlsx) file instead of CSV. Falls back to CSV when openpyxl not available.
        """
        import csv as _csv
        import json as _json

        p = Path(path)
        filt = set(filter or [])

        # Normalize results list and ensure TestData-like structure
        results = getattr(test_data, "results", None)
        if results is None:
            logger.warning("log_write: no results to log for %s", getattr(test_data, "id", ""))
            results = []

        # Helper to sanitize test point name to a safe column prefix
        def _prefix_name(name: str, index: int) -> str:
            if not name:
                name = f"tp{index}"
            # Replace whitespace and problematic chars
            return "".join(ch if ch.isalnum() or ch in ("-","_") else "_" for ch in str(name)).replace(" ", "_")

        # --- Unified header and row construction ---
        def _build_header_and_row():
            top_fields = ["timestamp", "serial", "id", "meta", "passed", "pcb_lot", "test_checksum"]
            # Always include top_fields; filters only apply to per-test attributes
            header_cols = top_fields.copy()

            # Build per-test columns. Allow duplicate names by appending index suffix when needed
            prefix_counts: dict[str, int] = {}
            for i, r in enumerate(results, start=1):
                tp = getattr(r, "tp", None)
                tp_name = getattr(tp, "name", None) if tp is not None else None
                prefix = _prefix_name(tp_name, i)
                # Ensure unique
                if prefix in prefix_counts:
                    prefix_counts[prefix] += 1
                    prefix = f"{prefix}_{prefix_counts[prefix]}"
                else:
                    prefix_counts[prefix] = 1

                # TestPoint attributes to expose
                tp_attrs = ["parameter", "frequency", "span", "limit_db", "direction"]
                for a in tp_attrs:
                    # filter only applies to per-test attributes
                    if a in filt:
                        continue
                    header_cols.append(f"{prefix}_{a}")

                # TestResult attributes
                res_attrs = ["passed", "min", "max", "failing", "samples"]
                for a in res_attrs:
                    # filter only applies to per-test attributes
                    if a in filt:
                        continue
                    header_cols.append(f"{prefix}_{a}")

            # Build row dict matching header columns
            row: dict = {}
            row["timestamp"] = getattr(test_data, "timestamp", None)
            row["serial"] = getattr(test_data, "serial", None)
            row["id"] = getattr(test_data, "id", None)
            row["meta"] = getattr(test_data, "meta", None)
            row["passed"] = getattr(test_data, "passed", None)
            row["pcb_lot"] = getattr(test_data, "pcb_lot", None)
            # Prefer checksum present on TestData; otherwise fall back to the lot's checksum
            try:
                checksum = getattr(test_data, "test_checksum", None)
            except Exception:
                checksum = None
            if checksum is None:
                try:
                    checksum = self.lot_checksum.get(self.current_lot_name) if getattr(self, "lot_checksum", None) is not None else None
                except Exception:
                    checksum = None
            row["test_checksum"] = checksum

            # Fill per-test values
            seen_prefixes: dict[str, int] = {}
            for i, r in enumerate(results, start=1):
                tp = getattr(r, "tp", None)
                tp_name = getattr(tp, "name", None) if tp is not None else None
                prefix = _prefix_name(tp_name, i)
                if prefix in seen_prefixes:
                    seen_prefixes[prefix] += 1
                    prefix = f"{prefix}_{seen_prefixes[prefix]}"
                else:
                    seen_prefixes[prefix] = 1

                if tp is not None:
                    tp_attrs = {
                        "parameter": getattr(tp, "parameter", None),
                        "frequency": getattr(tp, "frequency", None),
                        "span": getattr(tp, "span", None),
                        "limit_db": getattr(tp, "limit_db", None),
                        "direction": getattr(tp, "direction", None),
                    }
                else:
                    tp_attrs = {}
                for a, v in tp_attrs.items():
                    key = f"{prefix}_{a}"
                    row[key] = v

                res_attrs = {
                    "passed": getattr(r, "passed", None),
                    "min": getattr(r, "min", None),
                    "max": getattr(r, "max", None),
                    "failing": getattr(r, "failing", None),
                    "samples": getattr(r, "samples", None),
                }
                for a, v in res_attrs.items():
                    key = f"{prefix}_{a}"
                    if a == "failing":
                        row[key] = _json.dumps(v)
                    else:
                        row[key] = v

            # Ensure row has keys for header (top fields are preserved regardless of filter)
            for h in header_cols:
                row.setdefault(h, None)
            return header_cols, row

        def _read_existing_header(path_obj: Path, use_excel: bool) -> list | None:
            # Return existing header as list of column names, or None if not readable/existing
            if use_excel:
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(path_obj)
                    ws = wb.active
                    first_row = next(ws.iter_rows(min_row=1, max_row=1))
                    return [c.value for c in first_row]
                except Exception:
                    return None
            else:
                try:
                    if not path_obj.exists():
                        return None
                    with path_obj.open("r", encoding="utf-8", newline="") as f:
                        import csv as _csv_local
                        reader = _csv_local.reader(f)
                        return next(reader, None)
                except Exception:
                    return None

        # Build header and row
        header_cols, row = _build_header_and_row()

        # If excel requested, check openpyxl availability and fallback if missing
        if excel:
            try:
                import openpyxl  # noqa: F401
            except Exception:
                logger.warning("openpyxl not available; falling back to CSV for excel=True")
                excel = False

        # CSV branch
        if not excel:
            existing = _read_existing_header(p, use_excel=False)
            if existing:
                fieldnames = existing
                write_header = False
                mode = "a"
            else:
                fieldnames = header_cols
                write_header = True
                mode = "w"

            # Write CSV
            try:
                with p.open(mode, encoding="utf-8", newline="") as f:
                    writer = _csv.DictWriter(f, fieldnames=fieldnames)
                    if write_header:
                        writer.writeheader()
                    writer.writerow(row)
            except Exception:
                logger.exception("Failed writing CSV log row to %s", p)
            return

        # Excel branch
        # Ensure .xlsx suffix
        if p.suffix.lower() != ".xlsx":
            p = p.with_suffix(".xlsx")

        existing = _read_existing_header(p, use_excel=True)
        try:
            from openpyxl import Workbook, load_workbook
        except Exception:
            logger.exception("openpyxl import failed; cannot write excel file")
            # fallback to csv
            self.log_write(test_data, path, filter=filter, excel=False)
            return

        if existing:
            fieldnames = existing
        else:
            # create workbook and write header
            wb = Workbook()
            ws = wb.active
            fieldnames = header_cols
            ws.append(fieldnames)
            wb.save(p)

        # Append row in field order
        values = [row.get(fn, None) for fn in fieldnames]
        # Convert timestamp string to an actual datetime for Excel compatibility
        if "timestamp" in fieldnames:
            try:
                ti = fieldnames.index("timestamp")
                ts_val = values[ti]
                if isinstance(ts_val, str):
                    try:
                        # Prefer ISO format parsing
                        ts_parsed = datetime.fromisoformat(ts_val)
                    except Exception:
                        try:
                            ts_parsed = datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S")
                        except Exception:
                            ts_parsed = None
                    if ts_parsed is not None:
                        values[ti] = ts_parsed
            except Exception:
                # If anything goes wrong, leave the raw value as-is and continue
                logger.exception("Failed converting timestamp to datetime for Excel: %s", p)

        try:
            wb = load_workbook(p)
            ws = wb.active
            ws.append(values)
            wb.save(p)
        except Exception:
            logger.exception("Failed writing Excel log row to %s", p)



